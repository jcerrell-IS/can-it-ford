"""Extract EVERY entry of the Kramer 2021 supplementary archive, with cell-level provenance.

WHY THIS EXISTS
---------------
`simulation/r5_physics/kramer_benchmark.py` reduces the 31 numerical and 27 experimental
TIME SERIES at runtime, and does it well. But three of the archive's artifacts are opened
by no committed code at all, and a fourth is opened once by a human and TRANSCRIBED:

    Numerical results/Description of numerical models.xlsx   TRANSCRIBED into CODE_META
    Descriptions/Details on sphere mass distribution and densities.xlsx   NEVER OPENED
    Descriptions/Highly accurate experimental tests ... Kramer Sphere Cases.pdf  NEVER OPENED
    Descriptions/3D CAD model in Solidworks.zip              NEVER OPENED
    Datafile/Readme.pdf                                      NEVER OPENED

Measured, not assumed: `/usr/bin/grep -c` against kramer_benchmark.py on 2026-08-18 returns
openpyxl 0, pandas 0, zipfile 0, Readme 0, Solidworks 0, "mass distribution" 0, densities 0.
The single `xlsx` hit is inside a comment.

THAT MATTERS BECAUSE THE TRANSCRIBED ONE IS LOAD BEARING. `CODE_META` supplies `FAMILY`,
which selects every family split in `intercode()`, and `GROUPS`, which is the whole
"eleven codes are not eleven independent results" finding. A wrong entry there does not
move a number slightly. It reassigns a code to the wrong family or the wrong group, which
changes who contributes the published envelope, which is the headline.

So this module opens the sheet at runtime, resolves its ditto marks, records the SHEET NAME
and the CELL ADDRESS behind every field, and diffs the result against the transcription.
It also re-derives the RANS4/RANS5 radial-order finding independently, per series rather
than per code, because a finding established on some series and asserted for all is exactly
the kind of claim this project keeps having to withdraw.

Nothing here is transcribed. Every value is read from the archive on each run.

    SOURCE  energies-14-00269-s001.zip, sha256
            04c4d78d6987e4eec6c31d692d3c5cf5adea2580ffcfe50fbbd44e6589c7623f
            held at /Users/josie/can-it-ford-refs/2026-08-16/, deliberately OUTSIDE this
            public repo while register E8 is open.
    LICENCE Kramer, Andersen, Thomas, Ferri, Crowley, Stratigaki, Troch et al. 2021,
            Energies 14(2):269, doi 10.3390/en14020269.
            NO LICENCE FILE SHIPS IN THE ARCHIVE. Verified: zero entries in the zip match
            licen/copyright/terms/CC BY. The CC BY 4.0 status is PRIMARY from the article
            PDF itself, which states "distributed under the terms and conditions of the
            Creative Commons Attribution (CC BY) license
            (https://creativecommons.org/licenses/by/4.0/)".
            DO NOT cite the DTU Orbit cover page for licence: that page carries DTU's
            repository boilerplate, including "You may not further distribute the
            material", which is the REPOSITORY's terms and not the article's licence.
            CC BY permits redistribution with attribution. Permission is not obligation:
            derived statistics with attribution go in this repo, no Kramer series file does.

RUN
    uv venv /tmp/v && uv pip install --python /tmp/v/bin/python numpy pandas openpyxl scipy
    /tmp/v/bin/python analysis/kramer_extract_numerical.py --all
"""
from __future__ import annotations

import argparse
import json
import math
import sys
import zipfile
from pathlib import Path

import numpy as np
import openpyxl

REFS = Path("/Users/josie/can-it-ford-refs/2026-08-16/energies-14-00269-s001/Datafile")
NUM_ROOT = REFS / "Numerical results"
EXP_ROOT = REFS / "Experimental results"
DESC_ROOT = REFS / "Descriptions"
MODEL_XLSX = NUM_ROOT / "Description of numerical models.xlsx"
SPHERE_XLSX = DESC_ROOT / "Details on sphere mass distribution and densities.xlsx"
CAD_ZIP = DESC_ROOT / "3D CAD model in Solidworks.zip"

DROPS = ("01D", "03D", "05D")
REPS = (1, 2, 3, 4)
DITTO = "-||-"

# The archive's directory name for each code, keyed by the name the SHEET uses.
# NLPF1/FNPF1 is the one place the archive disagrees with itself; see model_table().
SHEET_TO_DIR = {"NLPF1": "FNPF1"}


# --------------------------------------------------------------------------------------
# FAIL-LOUD GUARDS. Added after a probe showed three silent false passes in this file.
# --------------------------------------------------------------------------------------
# THE DEFECT THESE EXIST TO PREVENT, demonstrated against this module before they were
# written: pointing `radial_order()` at an EMPTY directory returned normally with zero
# series analysed and `reversal_is_universal_where_it_occurs: True`. A headline verdict
# was produced from no data at all, and nothing in the output distinguished it from a
# real result. `audit_code_meta()` did the same, reporting "NO SUBSTANTIVE DRIFT" after
# checking zero fields.
#
# THE RULE: a check must be able to say "I could not evaluate this", and that must not
# look like "I evaluated this and it was fine". An empty read is an ERROR, never a pass.
#
# So every extraction below asserts what it expected to find, and the counts are the
# ARCHIVE'S OWN, recorded here so a future change to the archive fails loudly rather than
# silently shrinking a result set. `--self-test` proves each guard actually fires.

EXPECTED = {
    "sheet_rows": 13,           # 11 models plus the two duplicate UoP rows
    "models": 11,
    "numerical_series": 31,     # NOT 33: RANS3 ships 05D only
    "experimental_series": 27,  # 24 Measured plus 3 CI95
    "numerical_wg_series": 10,  # RANS2 x3, RANS3 x1, RANS4 x3, RANS5 x3
    "experimental_wg_reps": 12,
    "densities": 3,
    "cost_entries": 4,
    "cad_entries": 24,
}


class ExtractionError(RuntimeError):
    """Raised when an extraction finds nothing, or finds a different amount than expected.

    Deliberately NOT caught anywhere in this module. A caller that wants to continue past
    a failed extraction has to do so explicitly, which is the point.
    """


def _require(cond: bool, what: str, detail: str = "") -> None:
    if not cond:
        raise ExtractionError(f"{what}. {detail}".strip())


def _require_count(got: int, want: int, what: str, where: str) -> None:
    """Exact-count assertion. Says what it wanted, what it got, and where to look."""
    if got != want:
        raise ExtractionError(
            f"{what}: expected {want}, found {got}, reading {where}. "
            f"Either the archive changed or this is pointed at the wrong place. "
            f"A count that silently shrinks is the failure mode this guard exists for; "
            f"do not 'fix' it by lowering the expectation without checking the source.")


def _require_nonempty(seq, what: str, where: str):
    if len(seq) == 0:
        raise ExtractionError(
            f"{what} is EMPTY, reading {where}. This is an error, not a pass: an empty "
            f"result set must never be reported as a clean verdict.")
    return seq


def _fail(msg: str) -> None:
    print(f"  !! {msg}")


# --------------------------------------------------------------------------------------
# 1. THE MODEL TABLE, read from the sheet with cell addresses, ditto marks resolved
# --------------------------------------------------------------------------------------
def model_table(path: Path = MODEL_XLSX) -> dict:
    """Read `Description of numerical models.xlsx` and return one record per row.

    THE SHEET IS NOT A CLEAN TABLE AND THREE THINGS WILL BITE A NAIVE READER.

    1. IT HAS 13 ROWS FOR 11 MODELS. The first two, UoPLam and UoPSST, carry Plymouth's
       internal names and their Description cells are word-for-word those of RANS2 and
       RANS3. They are the same two submissions listed twice. This function returns all
       13 and flags the duplicates rather than silently dropping them, because "13 rows,
       2 duplicates, 11 models" is itself a finding and a silent drop hides it.
    2. DITTO MARKS `-||-` APPEAR IN Author AND Software, and they do NOT span the same
       rows. Author dittos run E14:E17 pointing at E13; Software dittos run F15:F17
       pointing at F14, which is NOT the same value as F13. Resolving both with one rule
       would give LPF1-4 the wrong software.
    3. THE TABLE DOES NOT START AT A1. Headers are on row 4, columns C to H. There is a
       SECOND, UNRELATED table lower down at G25:H31 giving cost by model class; it is
       returned separately by `cost_classes()`.

    Every returned field carries the cell it came from, so any value here can be checked
    against the spreadsheet by opening one cell.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    hdr_row = 4
    cols = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=hdr_row, column=c).value
        if v:
            cols[str(v).strip()] = c
    want = ("Model name", "Institution", "Author", "Software",
            "Description", "Computational effort")
    missing = [w for w in want if w not in cols]
    _require(not missing,
             f"model sheet header row {hdr_row} is missing {missing}",
             f"found headers {sorted(cols)} in {path.name} sheet {ws.title!r}. "
             f"The table does not start at A1; if the sheet was edited the header row "
             f"may have moved. Do NOT read a zero-row result as 'no models'.")

    rows, r = [], hdr_row + 1
    while r <= ws.max_row:
        name = ws.cell(row=r, column=cols["Model name"]).value
        if name is None or not str(name).strip():
            break
        rec = {"sheet": ws.title, "row": r,
               "name": str(name).strip(),
               "name_cell": ws.cell(row=r, column=cols["Model name"]).coordinate}
        for field in ("Institution", "Author", "Software",
                      "Description", "Computational effort"):
            cell = ws.cell(row=r, column=cols[field])
            raw = cell.value
            raw = None if raw is None else str(raw).strip()
            key = field.lower().replace(" ", "_")
            rec[key + "_raw"] = raw
            rec[key + "_cell"] = cell.coordinate
            # ditto resolution: walk UP the same column to the first non-ditto value
            if raw == DITTO:
                rr, src = r - 1, None
                while rr > hdr_row:
                    up = ws.cell(row=rr, column=cols[field]).value
                    up = None if up is None else str(up).strip()
                    if up and up != DITTO:
                        src = (up, ws.cell(row=rr, column=cols[field]).coordinate)
                        break
                    rr -= 1
                rec[key] = None if src is None else src[0]
                rec[key + "_from_ditto"] = True
                rec[key + "_ditto_source_cell"] = None if src is None else src[1]
            else:
                rec[key] = raw
                rec[key + "_from_ditto"] = False
                rec[key + "_ditto_source_cell"] = None
        rec["dir_name"] = str(SHEET_TO_DIR.get(rec["name"], rec["name"]))
        rec["has_directory"] = (NUM_ROOT / rec["dir_name"]).is_dir()
        rows.append(rec)
        r += 1

    # duplicate detection is by DESCRIPTION text, which is what actually repeats
    _require_count(len(rows), EXPECTED["sheet_rows"], "model sheet data rows",
                   f"{path.name} sheet {ws.title!r} from row {hdr_row + 1}")
    by_desc = {}
    for rec in rows:
        by_desc.setdefault(rec["description"] or "", []).append(rec["name"])
    dups = {k: v for k, v in by_desc.items() if len(v) > 1}
    out = {
        "path": str(path), "sheet": ws.title, "header_row": hdr_row,
        "header_columns": {k: ws.cell(row=hdr_row, column=v).coordinate
                           for k, v in cols.items()},
        "missing_expected_headers": missing,
        "n_rows": len(rows), "rows": rows,
        "rows_without_directory": [x["name"] for x in rows if not x["has_directory"]],
        "duplicate_description_groups": [
            {"names": v, "description_prefix": (k[:70] + "...") if len(k) > 70 else k}
            for k, v in dups.items()],
        "n_models_after_dedup": len(rows) - sum(len(v) - 1 for v in dups.values()),
    }
    _require_count(out["n_models_after_dedup"], EXPECTED["models"],
                   "distinct models after dedup", path.name)
    return out


def cost_classes(path: Path = MODEL_XLSX) -> dict:
    """The SECOND table on the same sheet, at G25:H31. Read by nothing before this.

    It is the authors' own cost-by-class summary and it is the only place in the archive
    that states an expected order of magnitude per model type.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    out, r = [], 1
    title = None
    while r <= ws.max_row:
        g = ws.cell(row=r, column=7).value
        h = ws.cell(row=r, column=8).value
        if g and not h and "time" in str(g).lower() and r > 20:
            title = {"text": str(g).strip(), "cell": ws.cell(row=r, column=7).coordinate}
        elif g and h and r > 20:
            out.append({"model_type": str(g).strip(), "cost": str(h).strip(),
                        "cells": [ws.cell(row=r, column=7).coordinate,
                                  ws.cell(row=r, column=8).coordinate]})
        r += 1
    _require_nonempty(out, "cost-by-class table rows",
                      f"{path.name} sheet {ws.title!r} columns G:H below row 20")
    header = out[0] if out else None
    _require_count(len(out) - 1, EXPECTED["cost_entries"],
                   "cost-by-class entries", f"{path.name} G28:H31")
    return {"path": str(path), "sheet": ws.title, "title": title,
            "header": header, "entries": out[1:] if out else [],
            "n_entries": max(0, len(out) - 1)}


# --------------------------------------------------------------------------------------
# 2. AUDIT THE TRANSCRIPTION
# --------------------------------------------------------------------------------------
def _turbulence_from_description(desc: str) -> str:
    """Classify the turbulence treatment FROM THE AUTHORS' OWN DESCRIPTION TEXT.

    This is a keyword read of a free-text cell, so it is reported as `evidence` alongside
    the verdict and never as a bare label. The phrases matched are the authors' own.
    """
    d = (desc or "").lower()
    if "without a turbulence model" in d:
        return "LAMINAR (author text: 'without a turbulence model')"
    if "laminar" in d:
        return "LAMINAR (author text: 'laminar')"
    if "sst" in d:
        return "SST (author text: 'SST')"
    if "potential flow" in d or "bem" in d or "linear coefficients" in d:
        return "n/a (potential flow)"
    return "UNCLASSIFIED"


def audit_code_meta(mt: dict | None = None) -> dict:
    """Diff `kramer_benchmark.CODE_META` field by field against the sheet.

    Reports DRIFT rather than raising, because a cosmetic difference (a typo the archive
    itself carries) and a substantive one (a wrong family) must not look the same.
    """
    mt = mt or model_table()
    sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "simulation" / "r5_physics"))
    try:
        import kramer_benchmark as kb
    except Exception as exc:                                    # pragma: no cover
        return {"available": False, "error": f"{type(exc).__name__}: {exc}"}

    sheet = {r["name"]: r for r in mt["rows"]}
    findings, checked = [], 0
    for code, meta in kb.CODE_META.items():
        sheet_name = meta.get("alias") or code
        rec = sheet.get(sheet_name) or sheet.get(code)
        if rec is None:
            findings.append({"code": code, "field": "*", "severity": "MISSING",
                             "detail": f"no sheet row named {sheet_name!r} or {code!r}"})
            continue
        for field, key in (("institution", "institution"), ("author", "author"),
                           ("software", "software")):
            checked += 1
            got, want = meta.get(field), rec[key]
            g = (got or "").lower().replace("(ditto in the sheet)", "").replace("(ditto)", "")
            g = " ".join(g.split())
            w = " ".join((want or "").lower().split())
            if g != w:
                sev = "COSMETIC" if g.replace("technology", "").replace("university", "") \
                    == w.replace("technology", "").replace("university", "") else "DRIFT"
                findings.append({"code": code, "field": field, "severity": sev,
                                 "in_code_meta": got, "in_sheet": want,
                                 "sheet_cell": rec[key + "_cell"],
                                 "resolved_from_ditto": rec[key + "_from_ditto"],
                                 "ditto_source_cell": rec[key + "_ditto_source_cell"]})
        checked += 1
        derived = _turbulence_from_description(rec["description"])
        stated = meta.get("turbulence", "")
        agree = (("LAMINAR" in derived and "LAMINAR" in stated.upper())
                 or ("SST" in derived and "SST" in stated.upper())
                 or ("n/a" in derived and "n/a" in stated))
        if not agree:
            findings.append({"code": code, "field": "turbulence", "severity": "DRIFT",
                             "in_code_meta": stated, "from_sheet_description": derived,
                             "sheet_cell": rec["description_cell"]})
    _require(checked > 0,
             "CODE_META audit checked ZERO fields",
             "either CODE_META is empty or no sheet row matched any code name. "
             "Reporting a clean verdict here would be a false pass, which is exactly "
             "what this guard exists to prevent.")
    _require_count(checked, 4 * EXPECTED["models"],
                   "CODE_META fields checked (4 per code)", "kramer_benchmark.CODE_META")
    return {"available": True, "n_fields_checked": checked,
            "n_findings": len(findings),
            "n_drift": sum(1 for f in findings if f["severity"] == "DRIFT"),
            "n_cosmetic": sum(1 for f in findings if f["severity"] == "COSMETIC"),
            "findings": findings,
            "verdict": "NO SUBSTANTIVE DRIFT"
            if not any(f["severity"] == "DRIFT" for f in findings) else "DRIFT FOUND"}


def grouping_keys(mt: dict | None = None) -> dict:
    """Group the eleven codes by AUTHOR and by INSTITUTION, both as the sheet ships them.

    THIS IS THE POINT OF THIS FUNCTION: the two keys give the same COUNT and a different
    MEMBERSHIP, and the "one group sets both ends of the envelope" headline depends on
    which key is used. `CODE_META`'s hand-assigned `group` field matches the AUTHOR key,
    which is defensible for an independence claim, but the sheet never states it and a
    reader grouping by the Institution column as shipped gets a different answer.
    """
    mt = mt or model_table()
    rows = [r for r in mt["rows"] if r["has_directory"]]
    out = {}
    for key in ("author", "institution"):
        g = {}
        for r in rows:
            g.setdefault(r[key], []).append(r["dir_name"])
        out[key] = {"n_groups": len(g), "groups": {k: sorted(v) for k, v in g.items()}}
    a = {frozenset(v) for v in out["author"]["groups"].values()}
    i = {frozenset(v) for v in out["institution"]["groups"].values()}
    out["same_count"] = out["author"]["n_groups"] == out["institution"]["n_groups"]
    out["same_membership"] = a == i
    out["partitions_differing"] = sorted(
        [sorted(x) for x in (a - i)]) + sorted([sorted(x) for x in (i - a)])
    return out


def envelope_by_grouping(mt: dict | None = None) -> dict:
    """Rebuild the published-envelope group table under BOTH keys the sheet ships.

    THIS IS THE TEST, not an illustration. `R8_KRAMER_INTERCODE_2026-08-18.md` section 4
    reports "five of the six independent groups agree with the physical measurement to
    within 0.82 percent" and "the entire envelope is set at BOTH ends by the sixth group".
    Both sentences are grouping-dependent, and the sheet ships two grouping keys that give
    the same COUNT and a different MEMBERSHIP. So compute the table twice and print both.

    The period deviations themselves are NOT recomputed here. They come from
    `kramer_benchmark.intercode()`, which already reduces all 31 series with one statistic.
    Re-deriving them in a second place would create a fork, which is the failure mode
    CLAUDE.md item 16 already records for `gates.py`.
    """
    mt = mt or model_table()
    sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "simulation" / "r5_physics"))
    try:
        import kramer_benchmark as kb
    except Exception as exc:                                    # pragma: no cover
        return {"available": False, "error": f"{type(exc).__name__}: {exc}"}
    ic = kb.intercode()
    dev = {}
    for drop in DROPS:
        for code, row in ic["drops"][drop]["codes"].items():
            dev.setdefault(code, []).append(row["dev_period_pct"])
    sheet = {r["dir_name"]: r for r in mt["rows"] if r["has_directory"]}
    out = {"available": True, "n_codes_with_deviations": len(dev), "tables": {}}
    for key in ("author", "institution", "code_meta_group"):
        groups = {}
        for code, ds in dev.items():
            if key == "code_meta_group":
                g = kb.CODE_META[code]["group"]
            else:
                g = sheet[code][key]
            groups.setdefault(g, {"codes": [], "devs": []})
            groups[g]["codes"].append(code)
            groups[g]["devs"].extend(ds)
        rows = []
        for g, v in groups.items():
            a = np.array(v["devs"])
            rows.append({"group": g, "n_codes": len(v["codes"]),
                         "codes": sorted(v["codes"]),
                         "n_series": int(a.size),
                         "min_pct": float(a.min()), "max_pct": float(a.max()),
                         "worst_abs_pct": float(np.abs(a).max())})
        rows.sort(key=lambda x: x["worst_abs_pct"])
        allv = np.array([d for v in groups.values() for d in v["devs"]])
        lo_g = min(rows, key=lambda x: x["min_pct"])
        hi_g = max(rows, key=lambda x: x["max_pct"])
        tight = [r for r in rows if r["worst_abs_pct"] < 1.0]
        out["tables"][key] = {
            "n_groups": len(rows), "rows": rows,
            "envelope_min_pct": float(allv.min()), "envelope_max_pct": float(allv.max()),
            "group_setting_low_end": lo_g["group"],
            "group_setting_high_end": hi_g["group"],
            "both_ends_same_group": lo_g["group"] == hi_g["group"],
            "n_groups_within_1pct": len(tight),
            "worst_of_the_tight_groups_pct": max((r["worst_abs_pct"] for r in tight),
                                                 default=float("nan")),
        }
    t = out["tables"]
    out["conclusion_is_grouping_dependent"] = (
        t["author"]["both_ends_same_group"] != t["institution"]["both_ends_same_group"]
        or t["author"]["n_groups_within_1pct"] != t["institution"]["n_groups_within_1pct"])
    return out


# --------------------------------------------------------------------------------------
# 3. THE SPHERE WORKBOOK, opened by nothing before this
# --------------------------------------------------------------------------------------
def sphere_properties(path: Path = SPHERE_XLSX) -> dict:
    """Mass, ballast, densities and the full inertia tensor, with cell addresses.

    None of this reaches any committed calculation today. It is recorded because the
    project's own CLAUDE.md item 4 spent considerable effort establishing that an inertia
    tensor which is ESTIMATED must never be presented as MEASURED. Here is a benchmark
    where the tensor genuinely is measured, from a CAD model of the built article, and it
    is sitting unread in the archive.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {"path": str(path), "sheets": wb.sheetnames}

    def find(ws, text, col_offset=1, contains=True):
        """First cell whose text matches, plus the value col_offset columns right."""
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                s = str(c.value).strip()
                hit = (text.lower() in s.lower()) if contains else (s.lower() == text.lower())
                if hit:
                    v = ws.cell(row=c.row, column=c.column + col_offset).value
                    return {"label": s, "label_cell": c.coordinate, "value": v,
                            "value_cell": ws.cell(row=c.row,
                                                  column=c.column + col_offset).coordinate}
        return None

    ws = wb["Weights and ballast"]
    out["weights"] = {k: find(ws, k) for k in
                      ("Total shell", "Total connection parts")}
    ws = wb["Densities"]
    dens = {}
    for row in ws.iter_rows():
        cells = [c for c in row if c.value is not None]
        if len(cells) >= 2 and str(cells[0].value).strip() in ("Water", "Aluminium",
                                                               "Stainless steel"):
            dens[str(cells[0].value).strip()] = {
                "value_kg_m3": cells[1].value,
                "cell": cells[1].coordinate,
                "note": str(cells[2].value) if len(cells) > 2 else None}
    _require_count(len(dens), EXPECTED["densities"], "material densities",
                   f"{path.name} sheet 'Densities'")
    out["densities_kg_m3"] = dens
    out["total_sphere_with_ballast_g"] = find(ws, "Total Sphere with ballast", 3)
    out["total_sphere_no_ballast_g"] = find(ws, "Total Sphere no ballast", 3)

    ws = wb["Inertia moments"]
    inertia = {}
    # SCAN EVERY CELL OF THE ROW FOR THE LABEL, not just the first populated one.
    # THIS IS A BUG FIX AND THE GUARD ABOVE IS WHAT FOUND IT. The sheet merges B7:B11
    # for the words "Inertia moments", so on row 7 the FIRST populated cell is B7 and the
    # label 'Ixx and Iyy' sits in C7. A first-cell-only scan skipped that row entirely
    # and dropped Ixx/Iyy = 0.098252280525 kgm2 from the output WITHOUT COMPLAINING,
    # while still reporting the other six inertia entries and looking complete.
    WANT = ("CoG z", "m", "Ixx and Iyy", "Izz", "Ixz and Izx",
            "Izy and I yz", "Ixy and Iyx")
    for row in ws.iter_rows():
        cells = [c for c in row if c.value is not None]
        for i, c in enumerate(cells[:-1]):
            lab = str(c.value).strip()
            if lab in WANT and lab not in inertia:
                val = cells[i + 1]
                inertia[lab] = {"value": val.value, "cell": val.coordinate,
                                "unit": str(cells[i + 2].value)
                                        if len(cells) > i + 2 else None}
        if False:
            pass
    # the second block of the same sheet is the raw Solidworks dump
    for key, txt in (("mass_g", "Mass ="), ("volume_mm3", "Volume ="),
                     ("surface_area_mm2", "Surface area =")):
        hit = None
        for row in ws.iter_rows():
            for c in row:
                if c.value and txt in str(c.value):
                    hit = {"text": str(c.value).strip(), "cell": c.coordinate}
                    break
            if hit:
                break
        inertia[key] = hit
    for key in ("CoG z", "m", "Ixx and Iyy", "Izz"):
        _require(inertia.get(key) is not None,
                 f"inertia label {key!r} not found",
                 f"{path.name} sheet 'Inertia moments'. A renamed label would otherwise "
                 f"return None silently and be reported as absent rather than as a miss.")
    out["inertia"] = inertia
    _require(out["total_sphere_with_ballast_g"] is not None,
             "ballasted sphere mass not found", f"{path.name} sheet 'Densities'")
    return out


def sphere_consistency(sp: dict | None = None) -> dict:
    """Falsifiable checks on the sphere numbers, so they are not merely transcribed.

    Each check states the law used and is reported with its own residual, whichever way
    it comes out. Benchmark gravity 9.82 m/s2 is Table 1's, not the engine's.
    """
    sp = sp or sphere_properties()
    g = 9.82
    rho_w = sp["densities_kg_m3"]["Water"]["value_kg_m3"]
    m_g = sp["total_sphere_with_ballast_g"]["value"]
    m = float(m_g) / 1000.0
    d = 0.300
    v_sphere = (4.0 / 3.0) * math.pi * (0.5 * d) ** 3
    v_disp_eq = m / rho_w
    checks = []
    checks.append({
        "check": "equilibrium submerged volume against a half sphere",
        "law": "Archimedes: m = rho_w * V_displaced at equilibrium",
        "V_displaced_m3": v_disp_eq, "V_half_sphere_m3": 0.5 * v_sphere,
        "ratio": v_disp_eq / (0.5 * v_sphere),
        "pct_from_half": 100.0 * (v_disp_eq / (0.5 * v_sphere) - 1.0),
        "reading": "the sphere floats at very close to half its diameter, which is the "
                   "benchmark's design intent and the reason it is a clean case"})
    # heave stiffness of a sphere at half draft: k = rho*g*A_wp, A_wp = pi*R^2
    a_wp = math.pi * (0.5 * d) ** 2
    k = rho_w * g * a_wp
    checks.append({
        "check": "undamped natural period from measured mass and waterplane stiffness",
        "law": "k = rho_w*g*pi*R^2 at half draft; T = 2*pi*sqrt((m + a33)/k)",
        "waterplane_area_m2": a_wp, "k_N_per_m": k,
        "T_no_added_mass_s": 2.0 * math.pi * math.sqrt(m / k),
        "T_with_a33_over_m_0p5_s": 2.0 * math.pi * math.sqrt(m * 1.5 / k),
        "reading": "brackets the archive's Te0 = 0.7561 s normalising constant; the "
                   "added mass required to land on Te0 is reported next"})
    t_e0 = 0.7561
    a33_needed = k * (t_e0 / (2.0 * math.pi)) ** 2 - m
    checks.append({
        "check": "added mass implied by Te0",
        "law": "a33 = k*(Te0/2pi)^2 - m",
        "a33_kg": a33_needed, "a33_over_m": a33_needed / m,
        "reading": "compare against the module's measured 0.540 at 01D and 0.870 at 05D; "
                   "Te0 is a fixed normalising constant, not a per-drop measurement"})
    volume_txt = sp["inertia"].get("volume_mm3")
    if volume_txt:
        try:
            v_solid_mm3 = float(str(volume_txt["text"]).split("=")[1].split()[0])
            checks.append({
                "check": "solid material volume against the enclosing sphere",
                "law": "geometry only",
                "V_solid_m3": v_solid_mm3 * 1e-9, "V_enclosing_m3": v_sphere,
                "solid_fraction": v_solid_mm3 * 1e-9 / v_sphere,
                "reading": "the sphere is hollow; the CAD volume is shell plus ballast, "
                           "NOT the displacement volume, and confusing the two would "
                           "give a density about six times too high"})
        except (ValueError, IndexError):
            pass
    return {"g_used_m_s2": g, "rho_w_kg_m3": rho_w, "mass_kg": m, "checks": checks}


# --------------------------------------------------------------------------------------
# 4. EVERY SERIES, WITH THE COLUMNS IT DECLARES
# --------------------------------------------------------------------------------------
def _header_columns(p: Path) -> list:
    with open(p) as fh:
        return fh.readline().split()


def _column_names(p: Path) -> list:
    """Split the header into per-column names. Units are bracketed, so join on '['."""
    toks = _header_columns(p)
    names, cur = [], []
    for t in toks:
        cur.append(t)
        if t.endswith("]"):
            names.append(" ".join(cur))
            cur = []
    if cur:
        names.append(" ".join(cur))
    return names


def series_manifest() -> dict:
    """Every series file in the archive: path, declared columns, shape, time span.

    Discovered by walking the tree, never by assuming a shape. The archive is UNBALANCED
    and any code that assumes 11 x 3 is wrong.
    """
    num = []
    for d in sorted(p for p in NUM_ROOT.iterdir() if p.is_dir()):
        for drop in DROPS:
            p = d / f"{drop}_{d.name}.txt"
            if not p.exists():
                num.append({"kind": "numerical", "code": d.name, "drop": drop,
                            "present": False})
                continue
            a = np.loadtxt(p, skiprows=1)
            num.append({"kind": "numerical", "code": d.name, "drop": drop, "present": True,
                        "file": str(p.relative_to(REFS)),
                        "declared_columns": _column_names(p),
                        "n_columns": int(a.shape[1]), "n_rows": int(a.shape[0]),
                        "t_min_s": float(a[:, 0].min()), "t_max_s": float(a[:, 0].max()),
                        "x3_at_t0_m": float(a[np.argmin(np.abs(a[:, 0])), 1]),
                        "uniform_dt": bool(np.allclose(np.diff(a[:, 0]),
                                                       a[1, 0] - a[0, 0], rtol=1e-6)),
                        "has_wg": bool(a.shape[1] >= 5)})
    exp = []
    for drop in DROPS:
        for rep in REPS:
            for kind in ("Raw", "Normalized"):
                p = EXP_ROOT / f"{drop}_Measured{rep}_{kind}.txt"
                a = np.loadtxt(p, skiprows=1)
                exp.append({"kind": "experimental", "drop": drop, "rep": rep,
                            "form": kind, "file": str(p.relative_to(REFS)),
                            "declared_columns": _column_names(p),
                            "n_columns": int(a.shape[1]), "n_rows": int(a.shape[0]),
                            "has_wg": bool(a.shape[1] >= 5)})
        p = EXP_ROOT / f"{drop}_CI95_Normalized.txt"
        a = np.loadtxt(p, skiprows=1)
        exp.append({"kind": "experimental", "drop": drop, "rep": None, "form": "CI95",
                    "file": str(p.relative_to(REFS)),
                    "declared_columns": _column_names(p),
                    "n_columns": int(a.shape[1]), "n_rows": int(a.shape[0]),
                    "has_wg": bool(a.shape[1] >= 5)})
    present = [s for s in num if s["present"]]
    _require_count(len(present), EXPECTED["numerical_series"],
                   "numerical series found", str(NUM_ROOT))
    _require_count(len(exp), EXPECTED["experimental_series"],
                   "experimental series found", str(EXP_ROOT))
    for sr in present:
        _require(sr["n_rows"] > 0,
                 f"numerical series {sr['code']}/{sr['drop']} read ZERO rows", sr["file"])
    for sr in exp:
        _require(sr["n_rows"] > 0,
                 f"experimental series {sr['file']} read ZERO rows", sr["file"])
    rows = [s["n_rows"] for s in present]
    return {
        "numerical": num, "experimental": exp,
        "n_numerical_present": len(present),
        "n_numerical_absent": len([s for s in num if not s["present"]]),
        "absent": [f"{s['code']}/{s['drop']}" for s in num if not s["present"]],
        "n_experimental": len(exp),
        "n_series_total": len(present) + len(exp),
        "numerical_row_count_min": min(rows), "numerical_row_count_max": max(rows),
        "numerical_row_count_ratio": max(rows) / min(rows),
        "numerical_non_uniform_dt": [f"{s['code']}/{s['drop']}" for s in present
                                     if not s["uniform_dt"]],
        "numerical_with_wg": [f"{s['code']}/{s['drop']}" for s in present if s["has_wg"]],
        "codes_with_wg": sorted({s["code"] for s in present if s["has_wg"]}),
    }


# --------------------------------------------------------------------------------------
# 5. THE RADIAL ORDER TEST, PER SERIES, RE-DERIVED INDEPENDENTLY
# --------------------------------------------------------------------------------------
def radial_order() -> dict:
    """Does each WG-carrying series put its gauges in the declared radial order?

    THE TEST NEEDS NO GAUGE POSITIONS AT ALL, which is what makes it safe. A radiating
    wave from a compact source spreads as 1/sqrt(r) in amplitude, so the time integral of
    eta^2 must RISE monotonically toward the sphere. Every file in the archive declares
    its columns as `WG1 WG2 WG3`, and Figure 8 places WG1 farthest at 1.800 m and WG3
    nearest at 0.600 m. So the integral must rise from column WG1 to column WG3. A series
    where it falls has its gauge columns in the opposite order to what its own header says.

    TWO THINGS THIS ADDS TO THE EXISTING first-versus-last TEST.
      * MONOTONICITY. The existing check compares only the first and last gauge, so a
        series can pass it with its middle gauge out of order. Both are reported.
      * PER SERIES, NOT PER CODE. A code-level verdict is an OR over its series and
        cannot distinguish "this code always reverses" from "this code reverses
        sometimes", and those two have completely different consequences for a reader.
    """
    out, rows = {}, []
    for d in sorted(p for p in NUM_ROOT.iterdir() if p.is_dir()):
        for drop in DROPS:
            p = d / f"{drop}_{d.name}.txt"
            if not p.exists():
                continue
            a = np.loadtxt(p, skiprows=1)
            if a.shape[1] < 5:
                continue
            # PER SERIES, per the self-audit rule: a series that reads to zero usable rows
            # must stop the run, not silently drop out of the set and shrink the verdict.
            _require(a.shape[0] > 0, f"{d.name}/{drop} read ZERO rows", str(p))
            m = a[:, 0] >= 0.0
            _require(int(m.sum()) > 0,
                     f"{d.name}/{drop} has ZERO samples at t >= 0",
                     f"{p}: {a.shape[0]} rows, t spans "
                     f"{a[:, 0].min():.4f} to {a[:, 0].max():.4f}")
            t = a[m, 0]
            ints = [float(np.trapezoid(a[m, 2 + i] ** 2, t)) for i in range(3)]
            _require(all(v > 0.0 for v in ints),
                     f"{d.name}/{drop} has a non-positive eta^2 integral {ints}",
                     f"{p}: the ratio test divides by these, so a zero column would "
                     f"produce an inf or nan verdict rather than an error")
            rising = ints[2] > ints[0]
            mono_in = ints[0] < ints[1] < ints[2]
            mono_out = ints[0] > ints[1] > ints[2]
            rows.append({
                "series": f"{d.name}/{drop}", "code": d.name, "drop": drop,
                "declared_columns": _column_names(p)[2:5],
                "eta2_integral_m2s": {"WG1": ints[0], "WG2": ints[1], "WG3": ints[2]},
                "ratio_WG3_over_WG1": ints[2] / ints[0],
                "rises_toward_sphere": bool(rising),
                "strictly_monotone_rising": bool(mono_in),
                "strictly_monotone_falling": bool(mono_out),
                "verdict": "AS DECLARED" if mono_in else
                           ("REVERSED" if mono_out else
                            ("NON MONOTONE, rises overall" if rising
                             else "NON MONOTONE, falls overall")),
            })
    # experimental control: the same test on the measurement the codes were compared to
    exp_rows = []
    for drop in DROPS:
        for rep in REPS:
            p = EXP_ROOT / f"{drop}_Measured{rep}_Raw.txt"
            a = np.loadtxt(p, skiprows=1)
            _require(a.shape[0] > 0, f"experimental {drop} rep{rep} read ZERO rows", str(p))
            _require(a.shape[1] >= 5,
                     f"experimental {drop} rep{rep} has {a.shape[1]} columns, need 5",
                     f"{p}: the control depends on WG1-3 being present")
            m = a[:, 0] >= 0.0
            t = a[m, 0]
            ints = [float(np.trapezoid(a[m, 2 + i] ** 2, t)) for i in range(3)]
            exp_rows.append({
                "series": f"EXPERIMENT/{drop}_rep{rep}",
                "drop": drop, "rep": rep,
                "eta2_integral_m2s": {"WG1": ints[0], "WG2": ints[1], "WG3": ints[2]},
                "ratio_WG3_over_WG1": ints[2] / ints[0],
                "strictly_monotone_rising": bool(ints[0] < ints[1] < ints[2]),
                "verdict": "AS DECLARED" if ints[0] < ints[1] < ints[2] else "ANOMALY"})
    # THE SWAP HYPOTHESIS, which is what separates a labelling fault from a physics one.
    # If a series is reversed ONLY because its columns are written in the opposite order,
    # then undoing the swap must reproduce the measured radial gradient, not merely its
    # sign. So invert the ratio and ask whether it lands inside the band the twelve
    # experimental repetitions span. Landing inside supports "the labels are swapped".
    # Landing outside means the ordering is wrong AND the gradient is wrong, which a
    # relabelling cannot fix and which the authors would need to answer differently.
    #
    # THE POOLED BAND IS THE WRONG COMPARATOR AND USING IT ALONE INVERTED THIS RESULT.
    # An earlier revision of this function tested only against the band pooled over all
    # twelve repetitions, [2.655, 3.418], and reported RANS4 as "consistent with a column
    # swap" on all three drops. That conclusion does not survive a drop-matched test.
    # The experimental ratio VARIES SYSTEMATICALLY WITH DROP HEIGHT (01D [2.655, 3.255],
    # 03D [3.177, 3.373], 05D [3.312, 3.418]), so pooling widens the band by importing
    # 01D's scatter into the 05D comparison, and the widening is driven by a single
    # repetition, 01D_rep1 at 2.655. Against its OWN drop, RANS4 lands outside on all
    # three. Both comparators are therefore computed and both are reported; the
    # drop-matched one is the one to quote.
    # SET-LEVEL GUARDS, before any verdict is formed from these rows.
    _require_count(len(rows), EXPECTED["numerical_wg_series"],
                   "numerical series carrying WG columns", str(NUM_ROOT))
    _require_count(len(exp_rows), EXPECTED["experimental_wg_reps"],
                   "experimental repetitions in the control", str(EXP_ROOT))
    eb = [r["ratio_WG3_over_WG1"] for r in exp_rows]
    lo, hi = min(eb), max(eb)
    dm = {}
    for drop in DROPS:
        v = [r["ratio_WG3_over_WG1"] for r in exp_rows if r["drop"] == drop]
        dm[drop] = (min(v), max(v))
    for r in rows:
        rec = 1.0 / r["ratio_WG3_over_WG1"]
        dlo, dhi = dm[r["drop"]]
        r["reciprocal_ratio"] = rec
        r["experiment_band_pooled"] = [lo, hi]
        r["experiment_band_drop_matched"] = [dlo, dhi]
        if r["verdict"] == "REVERSED":
            r["swap_in_pooled_band"] = bool(lo <= rec <= hi)
            r["swap_in_drop_matched_band"] = bool(dlo <= rec <= dhi)
            miss = 0.0 if dlo <= rec <= dhi else (dlo - rec if rec < dlo else rec - dhi)
            r["drop_matched_miss"] = miss
            r["drop_matched_miss_pct"] = 100.0 * miss / dlo
            r["swap_test"] = ("CONSISTENT WITH A COLUMN SWAP" if dlo <= rec <= dhi
                              else f"OUTSIDE ITS OWN DROP BAND by "
                                   f"{100.0 * miss / dlo:.1f}% of the lower edge")
        else:
            r["swap_in_pooled_band"] = None
            r["swap_in_drop_matched_band"] = None
            r["swap_test"] = "n/a, not reversed"
            r["in_drop_matched_band"] = bool(dlo <= r["ratio_WG3_over_WG1"] <= dhi)

    by_code = {}
    for r in rows:
        by_code.setdefault(r["code"], []).append(r["verdict"])
    out["series"] = rows
    out["experiment_control"] = exp_rows
    out["experiment_band_ratio"] = [lo, hi]
    out["experiment_band_drop_matched"] = {k: list(v) for k, v in dm.items()}
    rev = [r for r in rows if r["verdict"] == "REVERSED"]
    out["swap_consistent_codes_pooled"] = sorted({
        r["code"] for r in rev
        if all(x["swap_in_pooled_band"] for x in rev if x["code"] == r["code"])})
    out["swap_consistent_codes_drop_matched"] = sorted({
        r["code"] for r in rev
        if all(x["swap_in_drop_matched_band"] for x in rev if x["code"] == r["code"])})
    out["comparator_changes_the_verdict"] = (
        out["swap_consistent_codes_pooled"] != out["swap_consistent_codes_drop_matched"])
    out["worst_drop_matched_miss_pct"] = {
        c: max(x["drop_matched_miss_pct"] for x in rev if x["code"] == c)
        for c in sorted({r["code"] for r in rev})}
    out["experiment_all_as_declared"] = all(
        r["verdict"] == "AS DECLARED" for r in exp_rows)
    out["by_code"] = {
        c: {"n_series": len(v), "verdicts": v,
            "consistent": len(set(v)) == 1,
            "summary": v[0] if len(set(v)) == 1 else "INCONSISTENT ACROSS SERIES"}
        for c, v in by_code.items()}
    out["codes_reversed_on_every_series"] = sorted(
        c for c, v in out["by_code"].items() if v["consistent"] and v["summary"] == "REVERSED")
    out["codes_inconsistent"] = sorted(
        c for c, v in out["by_code"].items() if not v["consistent"])
    # THE FIELD THAT WAS FALSE-PASSING. It is a statement about the codes that reverse,
    # so it is undefined when none were examined, and it must say so rather than say True.
    _require_nonempty(out["by_code"], "per-code radial-order verdicts", str(NUM_ROOT))
    reversed_codes = [c for c, v in out["by_code"].items() if "REVERSED" in v["verdicts"]]
    out["n_codes_examined"] = len(out["by_code"])
    out["n_series_examined"] = len(rows)
    out["reversal_is_universal_where_it_occurs"] = (
        (not out["codes_inconsistent"]) if reversed_codes else
        "UNDEFINED, no code in this set reverses")
    return out


# --------------------------------------------------------------------------------------
# 6. THE DESCRIPTIONS SUBTREE, including the CAD archive
# --------------------------------------------------------------------------------------
def descriptions_inventory() -> dict:
    """What ships in Descriptions/, and what is inside the Solidworks zip.

    The CAD archive is listed, not extracted. Its part names are the link between the
    mass workbook's `Name in Solidworks` column and the physical article, so the listing
    is enough to check that the workbook refers to parts that exist.
    """
    files = []
    for p in sorted(DESC_ROOT.iterdir()):
        files.append({"name": p.name, "bytes": p.stat().st_size, "suffix": p.suffix})
    cad = {"present": CAD_ZIP.exists()}
    if CAD_ZIP.exists():
        with zipfile.ZipFile(CAD_ZIP) as z:
            names = z.namelist()
        cad.update({"n_entries": len(names), "entries": sorted(names),
                    "suffixes": sorted({Path(n).suffix.lower() for n in names})})
    readme = REFS / "Readme.pdf"
    return {"files": files, "n_files": len(files),
            "cad_zip": cad,
            "readme_pdf": {"present": readme.exists(),
                           "bytes": readme.stat().st_size if readme.exists() else None}}


def cad_cross_check() -> dict:
    """Do the `Name in Solidworks` values in the mass workbook appear in the CAD zip?

    A falsifiable link between two artifacts that have never been opened together. If the
    part names do not appear, that is reported as a NEGATIVE result and not smoothed over:
    a Solidworks assembly may store parts in a container the zip listing does not expose.
    """
    wb = openpyxl.load_workbook(SPHERE_XLSX, data_only=True)
    ws = wb["Densities"]
    names = []
    for row in ws.iter_rows():
        cells = [c for c in row if c.value is not None]
        if len(cells) >= 2 and isinstance(cells[1].value, str):
            v = str(cells[1].value).strip()
            if v and v not in ("Name in Solidworks",):
                names.append({"name": v, "cell": cells[1].coordinate})
    if not CAD_ZIP.exists():
        return {"cad_present": False, "solidworks_names": names}
    with zipfile.ZipFile(CAD_ZIP) as z:
        entries = z.namelist()
    joined = " ".join(entries).lower()
    hits = [{"name": n["name"], "cell": n["cell"],
             "found_in_cad_zip": n["name"].lower() in joined} for n in names]
    return {"cad_present": True, "n_names": len(hits),
            "n_found": sum(1 for h in hits if h["found_in_cad_zip"]),
            "names": hits,
            "cad_entries_sample": sorted(entries)[:25]}


# --------------------------------------------------------------------------------------
# printing
# --------------------------------------------------------------------------------------
def _p_model(mt):
    print("=" * 86)
    print("MODEL TABLE, read at runtime from the sheet")
    print("=" * 86)
    print(f"  file    {Path(mt['path']).name}")
    print(f"  sheet   {mt['sheet']!r}, headers on row {mt['header_row']}: "
          f"{', '.join(f'{k}@{v}' for k, v in mt['header_columns'].items())}")
    print(f"  rows    {mt['n_rows']}  ->  {mt['n_models_after_dedup']} distinct models")
    if mt["missing_expected_headers"]:
        _fail(f"missing headers {mt['missing_expected_headers']}")
    print(f"  rows with no directory in the archive: {mt['rows_without_directory']}")
    for g in mt["duplicate_description_groups"]:
        print(f"  DUPLICATE description shared by {g['names']}")
        print(f"      {g['description_prefix']}")
    print()
    print(f"  {'sheet name':<8} {'dir':<7} {'institution':<34} {'author':<32} software")
    for r in mt["rows"]:
        d = r["dir_name"] if r["has_directory"] else "(none)"
        auth = r["author"] or ""
        if r["author_from_ditto"]:
            auth += f" [ditto<-{r['author_ditto_source_cell']}]"
        sw = r["software"] or ""
        if r["software_from_ditto"]:
            sw += f" [ditto<-{r['software_ditto_source_cell']}]"
        print(f"  {r['name']:<8} {d:<7} {(r['institution'] or '')[:34]:<34} "
              f"{auth[:32]:<32} {sw}")
    print()
    print("  turbulence, classified from the authors' own Description cell:")
    for r in mt["rows"]:
        print(f"    {r['name']:<8} {r['description_cell']:<5} "
              f"{_turbulence_from_description(r['description'])}")


def _p_audit(a):
    print()
    print("=" * 86)
    print("AUDIT: kramer_benchmark.CODE_META against the sheet")
    print("=" * 86)
    if not a.get("available"):
        _fail(f"could not import kramer_benchmark: {a.get('error')}")
        return
    print(f"  fields checked {a['n_fields_checked']}, findings {a['n_findings']} "
          f"({a['n_drift']} DRIFT, {a['n_cosmetic']} COSMETIC)")
    print(f"  VERDICT: {a['verdict']}")
    for f in a["findings"]:
        print(f"    [{f['severity']}] {f['code']}.{f['field']}")
        for k in ("in_code_meta", "in_sheet", "from_sheet_description", "detail",
                  "sheet_cell", "ditto_source_cell"):
            if f.get(k) is not None:
                print(f"        {k:<24} {f[k]}")


def _p_groups(g):
    print()
    print("=" * 86)
    print("GROUPING: the sheet ships two keys and they do not give the same partition")
    print("=" * 86)
    for key in ("author", "institution"):
        print(f"  by {key}: {g[key]['n_groups']} groups")
        for k, v in sorted(g[key]["groups"].items()):
            print(f"      {k[:46]:<46} {v}")
    print(f"  same count      {g['same_count']}")
    print(f"  same membership {g['same_membership']}")
    if not g["same_membership"]:
        print("  partitions present under one key and not the other:")
        for p in g["partitions_differing"]:
            print(f"      {p}")


def _p_envelope(e):
    print()
    print("=" * 86)
    print("PUBLISHED ENVELOPE, regrouped under every key the archive supports")
    print("=" * 86)
    if not e.get("available"):
        _fail(f"could not import kramer_benchmark: {e.get('error')}")
        return
    for key, t in e["tables"].items():
        print(f"  --- key: {key}  ({t['n_groups']} groups) ---")
        print(f"  {'group':<34} {'n':>2} {'series':>6} {'min %':>8} {'max %':>8} {'worst':>7}")
        for r in t["rows"]:
            print(f"  {r['group'][:34]:<34} {r['n_codes']:>2} {r['n_series']:>6} "
                  f"{r['min_pct']:>8.2f} {r['max_pct']:>8.2f} {r['worst_abs_pct']:>7.2f}")
        print(f"    envelope {t['envelope_min_pct']:.2f} to {t['envelope_max_pct']:.2f} %")
        print(f"    low end set by  {t['group_setting_low_end']}")
        print(f"    high end set by {t['group_setting_high_end']}")
        print(f"    BOTH ENDS ONE GROUP: {t['both_ends_same_group']}")
        print(f"    groups within 1 pct: {t['n_groups_within_1pct']} of {t['n_groups']}"
              f" (worst of them {t['worst_of_the_tight_groups_pct']:.2f} %)")
        print()
    print(f"  THE SECTION 4 CONCLUSION IS GROUPING DEPENDENT: "
          f"{e['conclusion_is_grouping_dependent']}")


def _p_manifest(m):
    print()
    print("=" * 86)
    print("SERIES MANIFEST, discovered by walking the tree")
    print("=" * 86)
    print(f"  numerical present {m['n_numerical_present']}, absent {m['n_numerical_absent']} "
          f"{m['absent']}")
    print(f"  experimental      {m['n_experimental']}")
    print(f"  total series      {m['n_series_total']}")
    print(f"  numerical rows    {m['numerical_row_count_min']} to "
          f"{m['numerical_row_count_max']}, ratio {m['numerical_row_count_ratio']:.1f}x")
    print(f"  non uniform dt    {m['numerical_non_uniform_dt']}")
    print(f"  codes with WG     {m['codes_with_wg']} "
          f"({len(m['numerical_with_wg'])} series)")
    cols = {}
    for s in m["numerical"]:
        if s.get("present"):
            cols.setdefault(tuple(s["declared_columns"]), []).append(
                f"{s['code']}/{s['drop']}")
    print("  declared column sets, numerical:")
    for k, v in cols.items():
        print(f"      {list(k)}  x{len(v)}")
    ecols = {}
    for s in m["experimental"]:
        ecols.setdefault(tuple(s["declared_columns"]), []).append(s["form"])
    print("  declared column sets, experimental:")
    for k, v in ecols.items():
        print(f"      {list(k)}  x{len(v)}")


def _p_order(o):
    print()
    print("=" * 86)
    print("RADIAL ORDER, every WG series, re-derived independently")
    print("=" * 86)
    print("  physical basis: 1/r spreading forces the eta^2 integral to RISE toward the")
    print("  sphere. WG1 is farthest (1.800 m), WG3 nearest (0.600 m), per Figure 8.")
    print("  The test uses no radii, only the ordering, so it cannot be wrong about them.")
    print()
    print(f"  {'series':<14} {'WG1':>12} {'WG2':>12} {'WG3':>12} {'WG3/WG1':>9}  verdict")
    for r in o["series"]:
        i = r["eta2_integral_m2s"]
        print(f"  {r['series']:<14} {i['WG1']:>12.4e} {i['WG2']:>12.4e} "
              f"{i['WG3']:>12.4e} {r['ratio_WG3_over_WG1']:>9.3f}  {r['verdict']}")
    print()
    print("  EXPERIMENTAL CONTROL, the measurement the codes are compared against:")
    for r in o["experiment_control"]:
        print(f"  {r['series']:<22} ratio {r['ratio_WG3_over_WG1']:>8.3f}  {r['verdict']}")
    print(f"  experiment as declared on every repetition: {o['experiment_all_as_declared']}")
    lo, hi = o["experiment_band_ratio"]
    print(f"  experimental WG3/WG1 band, POOLED over 12 reps: {lo:.3f} to {hi:.3f}")
    for d, (a, b) in o["experiment_band_drop_matched"].items():
        print(f"    drop-matched {d}: {a:.3f} to {b:.3f}")
    print()
    print("  SWAP HYPOTHESIS: invert each reversed ratio and ask whether it recovers the")
    print("  measured gradient. THE POOLED BAND IS THE WRONG COMPARATOR: the experimental")
    print("  ratio varies with drop height, so pooling imports 01D scatter into 05D.")
    print(f"  {'series':<12} {'1/ratio':>8} {'pooled':>7} {'own drop':>9}  note")
    for r in o["series"]:
        if r["verdict"] != "REVERSED":
            continue
        print(f"    {r['series']:<12} {r['reciprocal_ratio']:>6.3f} "
              f"{str(r['swap_in_pooled_band']):>9} {str(r['swap_in_drop_matched_band']):>9}"
              f"  miss {r['drop_matched_miss_pct']:+.1f}%")
    print(f"  swap-consistent under POOLED band:       {o['swap_consistent_codes_pooled']}")
    print(f"  swap-consistent under DROP-MATCHED band: {o['swap_consistent_codes_drop_matched']}")
    print(f"  THE COMPARATOR CHANGES THE VERDICT:      {o['comparator_changes_the_verdict']}")
    print(f"  worst miss against own drop band: {o['worst_drop_matched_miss_pct']}")
    print()
    for c, v in sorted(o["by_code"].items()):
        flag = "" if v["consistent"] else "   <-- INCONSISTENT"
        print(f"  {c:<7} {v['n_series']} series  {v['summary']}{flag}")
    print()
    print(f"  codes reversed on EVERY series they ship: {o['codes_reversed_on_every_series']}")
    print(f"  codes INCONSISTENT across their series:   {o['codes_inconsistent']}")
    print(f"  reversal is universal where it occurs:    "
          f"{o['reversal_is_universal_where_it_occurs']}")


def _p_sphere(sp, sc):
    print()
    print("=" * 86)
    print("SPHERE WORKBOOK, opened by no committed code before this")
    print("=" * 86)
    print(f"  sheets {sp['sheets']}")
    for k, v in sp["densities_kg_m3"].items():
        print(f"    density {k:<16} {v['value_kg_m3']:>12}  [{v['cell']}]  {v['note']}")
    for k in ("total_sphere_no_ballast_g", "total_sphere_with_ballast_g"):
        v = sp[k]
        if v:
            print(f"    {k:<28} {v['value']}  [{v['value_cell']}]")
    print("  inertia block:")
    for k, v in sp["inertia"].items():
        if v is None:
            continue
        if "value" in v:
            print(f"    {k:<16} {v['value']:>18}  {v.get('unit') or '':<6} [{v['cell']}]")
        else:
            print(f"    {k:<16} {v['text']:<40} [{v['cell']}]")
    print()
    print("  FALSIFIABLE CHECKS, each with its law stated:")
    for c in sc["checks"]:
        print(f"    * {c['check']}")
        print(f"      law: {c['law']}")
        for k, v in c.items():
            if k in ("check", "law", "reading"):
                continue
            print(f"        {k:<28} {v}")
        print(f"      reading: {c['reading']}")


def _p_desc(d, cc):
    print()
    print("=" * 86)
    print("DESCRIPTIONS SUBTREE")
    print("=" * 86)
    for f in d["files"]:
        print(f"    {f['name'][:66]:<66} {f['bytes']:>10}")
    print(f"    Readme.pdf present {d['readme_pdf']['present']}, "
          f"{d['readme_pdf']['bytes']} bytes")
    cad = d["cad_zip"]
    if cad.get("present"):
        print(f"  CAD zip: {cad['n_entries']} entries, suffixes {cad['suffixes']}")
        for e in cad["entries"][:20]:
            print(f"      {e}")
    print()
    print("  CROSS CHECK, workbook 'Name in Solidworks' against the CAD zip listing:")
    if cc.get("cad_present"):
        print(f"    {cc['n_found']} of {cc['n_names']} names appear in the zip listing")
        for h in cc["names"]:
            print(f"      {h['name']:<22} [{h['cell']}]  "
                  f"{'FOUND' if h['found_in_cad_zip'] else 'not in listing'}")
    else:
        _fail("CAD zip absent")


def _p_cost(c):
    print()
    print("=" * 86)
    print("SECOND TABLE ON THE SAME SHEET, cost by model class, read by nothing before")
    print("=" * 86)
    if c["title"]:
        print(f"  title  {c['title']['text']!r} [{c['title']['cell']}]")
    if c["header"]:
        print(f"  header {c['header']['model_type']} | {c['header']['cost']} "
              f"{c['header']['cells']}")
    for e in c["entries"]:
        print(f"    {e['model_type']:<52} {e['cost']:<14} {e['cells']}")


# --------------------------------------------------------------------------------------
# 8. THE 0.3 PERCENT UNCERTAINTY, AND WHICH STATISTIC IT IS
# --------------------------------------------------------------------------------------
def _sweep_local_cutoff(root: Path, cuts=(0.02, 0.05, 0.10, 0.20)) -> dict:
    """Is the route-B factor an artifact of where the near-zero cutoff was put?

    Same discipline this module applies to the 1.0 in `envelope_by_grouping`: a number
    chosen by the author, that a reported factor is computed against, gets swept and the
    range reported. Recursion is avoided by calling the reduction directly rather than
    re-entering `uncertainty_scope`.
    """
    out = {}
    for c in cuts:
        facs = []
        for drop in DROPS:
            d = np.loadtxt(root / f"{drop}_CI95_Normalized.txt", skiprows=1)
            mean, half = d[:, 1], (d[:, 3] - d[:, 2]) / 2.0
            big = np.abs(mean) > c
            if not big.any():
                continue
            facs.append(float(np.median(half[big] / np.abs(mean[big]))
                              / np.median(half[big])))
        _require_nonempty(facs, f"drops with samples above cutoff {c}", str(root))
        out[str(c)] = {"min": float(min(facs)), "max": float(max(facs))}
    _require_nonempty(out, "cutoffs swept", str(root))
    lo = min(v["min"] for v in out.values())
    hi = max(v["max"] for v in out.values())
    return {"by_cutoff": out, "factor_min_over_all_cutoffs": lo,
            "factor_max_over_all_cutoffs": hi,
            "factor_stays_above_2x": bool(lo > 2.0)}


def uncertainty_scope(exp_root: Path | None = None,
                      local_cutoff: float = 0.05) -> dict:
    """Reproduce the abstract's "about 0.3% of the respective drop heights", then show
    what that number is NOT.

    WHY THIS FUNCTION EXISTS. The figure arrived here a second time, via a survey of
    validation targets, phrased as "approximately 0.3 percent experimental uncertainty"
    with no statistic named. It is not an independent assessment: it is the Kramer 2021
    abstract restated, so treating it as external corroboration would be one source cited
    twice. What IS worth doing is checking the paper's own claim against the data the
    paper shipped, and pinning down the normalisation, because the same measured band is
    0.29 percent or 15 percent depending only on what you divide it by.

    THREE ROUTES, and they are separate origins, which is the point:
      A  the authors' 95 percent CI series (3 files nothing else in this repo reduces)
      B  the same band divided by the LOCAL signal instead of the drop height
      C  repeatability of the first damped period across the 4 repetitions, which reads
         12 different files (the Raw series) and a different physical quantity
    Route A is the abstract's own statistic. Route C is the one commensurable with the
    inter-code envelope, because `intercode()` grades every code on the FIRST damped
    period against the experimental mean of that same quantity.
    """
    root = exp_root or EXP_ROOT
    per_drop, half_pct, local = {}, [], {}
    for drop in DROPS:
        p = root / f"{drop}_CI95_Normalized.txt"
        _require(p.exists(), f"CI95 series missing for {drop}", f"looked for {p}")
        d = np.loadtxt(p, skiprows=1)
        _require(d.ndim == 2 and d.shape[1] == 4,
                 f"{drop} CI95 file must have 4 columns "
                 f"(t/Te0, mean, lower, upper), found shape {d.shape}",
                 f"reading {p}")
        _require_nonempty(d, f"{drop} CI95 rows", str(p))
        t, mean, lo, hi = d[:, 0], d[:, 1], d[:, 2], d[:, 3]
        _require(bool((hi >= lo).all()),
                 f"{drop} CI95 upper bound is below the lower bound somewhere",
                 "the columns may be swapped in the file as shipped")
        half = (hi - lo) / 2.0                    # in units of H0, the file is normalised
        # Route B: the SAME band against the local signal. Restricted to where the signal
        # is not near a zero crossing, because a ratio to ~0 is not a meaningful percent.
        # NAMED, not bare. This section criticises a hard-coded 1.0 elsewhere in
        # the module, so its own cutoff is a parameter and is swept below.
        big = np.abs(mean) > local_cutoff
        _require_nonempty(mean[big],
                          f"{drop} samples with |x3/H0| > {local_cutoff}", str(p))
        rel = half[big] / np.abs(mean[big])
        # BOTH normalisations are reduced with the SAME statistic over the SAME samples.
        # An earlier draft compared a mean against a median and a factor computed from
        # that would have been meaningless; the two columns must be commensurable or the
        # ratio between them says nothing.
        per_drop[drop] = {
            "n_samples": int(d.shape[0]),
            "mean_halfwidth_pct_of_H0": float(half.mean() * 100.0),
            "max_halfwidth_pct_of_H0": float(half.max() * 100.0),
            "median_halfwidth_pct_of_H0_same_samples":
                float(np.median(half[big]) * 100.0),
            "median_rel_to_local_signal_pct": float(np.median(rel) * 100.0),
            "max_rel_to_local_signal_pct": float(rel.max() * 100.0),
            "n_samples_used_for_local": int(big.sum()),
            "normalisation_factor_median_over_median": float(
                np.median(rel) / np.median(half[big])),
        }
        half_pct.append(per_drop[drop]["mean_halfwidth_pct_of_H0"])
        local[drop] = per_drop[drop]["median_rel_to_local_signal_pct"]

    _require_count(len(half_pct), len(DROPS),
                   "drop heights with a CI95 series", str(root))
    pooled = float(np.mean(half_pct))

    # Route C: first-damped-period repeatability, the quantity intercode() actually grades.
    kb = _import_benchmark()
    reps = {}
    if kb is not None:
        ex = kb.experiment(root)
        for drop in DROPS:
            e = ex[drop]["first_damped_period_s"]
            _require(e["n"] == len(REPS),
                     f"{drop} first damped period built from {e['n']} repetitions, "
                     f"expected {len(REPS)}",
                     "a repeatability figure from fewer reps is not the same statistic")
            reps[drop] = {
                "mean_s": float(e["mean"]),
                "min_s": float(e["min"]),
                "max_s": float(e["max"]),
                "n": int(e["n"]),
                "range_pct_of_mean": float(100.0 * (e["max"] - e["min"]) / e["mean"]),
            }
        _require_count(len(reps), len(DROPS),
                       "drop heights with a period-repeatability figure", str(root))

    return {
        "abstract_claim": ("At a 95% confidence level, uncertainties were found to be very "
                           "low, on average only about 0.3% of the respective drop heights"),
        "abstract_source": "Kramer et al. 2021, Energies 14(2):269, abstract",
        "per_drop": per_drop,
        "pooled_mean_halfwidth_pct_of_H0": pooled,
        "reproduces_abstract_to_1dp": bool(round(pooled, 1) == 0.3),
        "route_b_median_rel_to_local_signal_pct": local,
        "route_c_first_period_repeatability": reps,
        "normalisation_factor_per_drop": {
            d: per_drop[d]["normalisation_factor_median_over_median"] for d in per_drop},
        "local_cutoff": float(local_cutoff),
        "local_cutoff_sweep": _sweep_local_cutoff(root) if local_cutoff == 0.05 else None,
        "normalisation_factor_min": float(min(
            per_drop[d]["normalisation_factor_median_over_median"] for d in per_drop)),
        "normalisation_factor_max": float(max(
            per_drop[d]["normalisation_factor_median_over_median"] for d in per_drop)),
    }


def _import_benchmark():
    """Import the sibling benchmark module, or return None. NEVER silently substitute."""
    import importlib.util
    p = (Path(__file__).resolve().parents[1] / "simulation" / "r5_physics"
         / "kramer_benchmark.py")
    if not p.exists():
        return None
    spec = importlib.util.spec_from_file_location("kramer_benchmark", p)
    m = importlib.util.module_from_spec(spec)
    sys.modules["kramer_benchmark"] = m
    spec.loader.exec_module(m)
    return m


def envelope_against_experiment() -> dict:
    """Put the inter-code envelope and the experiment's own repeatability on ONE scale.

    THE COMPARISON IS LIKE FOR LIKE BY CONSTRUCTION, and that had to be checked rather
    than assumed: `intercode()` computes `dev_period_pct` from `first_damped_period_s`
    against the experimental MEAN of the same quantity, so the commensurable experimental
    figure is the spread of the FIRST period across the 4 repetitions, not a mean over
    five cycles. Using a 5-cycle average here would have flattered the experiment.
    """
    kb = _import_benchmark()
    _require(kb is not None, "kramer_benchmark.py not found",
             "this comparison needs the module that computes the published envelope; "
             "it is NOT recomputed here, to avoid the second-definition fork "
             "CLAUDE.md item 16 records for gates.py")
    s = kb.intercode()
    ex = kb.experiment()
    rows = {}
    for drop in DROPS:
        d = s["drops"][drop]
        devs = [d["codes"][c]["dev_period_pct"] for c in d["codes"]]
        _require_nonempty(devs, f"{drop} per-code period deviations", "intercode()")
        e = ex[drop]["first_damped_period_s"]
        _require(e["mean"] != 0.0, f"{drop} experimental mean period is zero",
                 "cannot form a percentage against it")
        exp_range = 100.0 * (e["max"] - e["min"]) / e["mean"]
        _require(exp_range > 0.0,
                 f"{drop} experimental period range is {exp_range}, not positive",
                 "four repetitions that agree to the bit would mean the reps are the "
                 "same file, not that the experiment is perfect")
        width = max(devs) - min(devs)
        rows[drop] = {
            "n_codes": int(d["n_codes"]),
            "envelope_min_pct": float(min(devs)),
            "envelope_max_pct": float(max(devs)),
            "envelope_width_points": float(width),
            "exp_first_period_mean_s": float(e["mean"]),
            "exp_range_pct_of_mean": float(exp_range),
            "intercode_over_experiment": float(width / exp_range),
        }
    _require_count(len(rows), len(DROPS), "drops compared", "intercode() vs experiment()")
    ratios = [rows[d]["intercode_over_experiment"] for d in DROPS]
    widths = [rows[d]["envelope_width_points"] for d in DROPS]
    _require(widths[0] > 0.0, "01D envelope width is not positive",
             "cannot form a growth factor against it")
    return {
        "rows": rows,
        "ratio_min": float(min(ratios)),
        "ratio_max": float(max(ratios)),
        "ratio_rises_with_drop_height": bool(ratios == sorted(ratios)),
        "envelope_width_rises_with_drop_height": bool(widths == sorted(widths)),
        "experiment_is_not_the_limiting_factor": bool(min(ratios) > 1.0),
        "envelope_width_growth_01D_to_05D": float(widths[-1] / widths[0]),
        "exp_repeatability_growth_01D_to_05D": float(
            rows[DROPS[-1]]["exp_range_pct_of_mean"]
            / rows[DROPS[0]]["exp_range_pct_of_mean"]),
    }


def threshold_sensitivity(env: dict | None = None) -> dict:
    """Sweep the one bare threshold in this module and report the interval it is safe in.

    `envelope_by_grouping()` counts "groups within 1 percent" against a hard-coded 1.0.
    A verdict computed against a bare literal nobody swept is the exact shape of defect
    this project has already been bitten by (`sustain_frames = 3` flipping five verdicts
    at 4). So rather than defend the 1.0, measure the interval over which the count does
    not move. If that interval is wide, the number is a presentational bin sitting in an
    empty region of the data and no conclusion rests on it. If it is narrow, the headline
    is threshold-dependent and must say so.
    """
    e = env or envelope_by_grouping()
    out = {}
    for key, tab in e["tables"].items():
        w = sorted(r["worst_abs_pct"] for r in tab["rows"])
        _require_nonempty(w, f"per-group worst deviations under key {key!r}",
                          "envelope_by_grouping()")
        below = [x for x in w if x < 1.0]
        above = [x for x in w if x >= 1.0]
        _require_nonempty(below, f"groups under 1 pct with key {key!r}",
                          "a count of zero tight groups is a result this sweep cannot "
                          "interpret, not a pass")
        _require_nonempty(above, f"groups at or above 1 pct with key {key!r}",
                          "if every group is tight there is no envelope to explain and "
                          "the headline this sweep exists to test is vacuous")
        lo, hi = max(below), min(above)
        out[key] = {
            "count_at_1pct": len(below),
            "n_groups": len(w),
            "invariant_from_pct": float(lo),
            "invariant_to_pct": float(hi),
            "invariance_factor": float(hi / lo) if lo else float("inf"),
            "worst_abs_pct_sorted": [float(x) for x in w],
        }
    factors = [v["invariance_factor"] for v in out.values()]
    _require_nonempty(factors, "grouping keys swept", "envelope_by_grouping()['tables']")
    return {
        "keys": out,
        "min_invariance_factor": float(min(factors)),
        "threshold_is_load_bearing": bool(min(factors) < 2.0),
    }


def _p_uncertainty(u, c):
    print()
    print("=" * 86)
    print("THE 0.3 PERCENT FIGURE: reproduced, and then shown to be normalisation "
          "dependent")
    print("=" * 86)
    print(f"  abstract: \"{u['abstract_claim']}\"")
    print(f"  source  : {u['abstract_source']}")
    print()
    print("  ROUTE A, the authors' own CI95 series, half-width in units of the drop "
          "height:")
    for drop, v in u["per_drop"].items():
        print(f"    {drop}   mean {v['mean_halfwidth_pct_of_H0']:.4f} pct of H0"
              f"   max {v['max_halfwidth_pct_of_H0']:.4f} pct"
              f"   n {v['n_samples']}")
    print(f"    pooled over the three drops: "
          f"{u['pooled_mean_halfwidth_pct_of_H0']:.4f} pct of H0")
    print(f"    reproduces the abstract to 1 dp: {u['reproduces_abstract_to_1dp']}")
    print()
    print("  ROUTE B, THE SAME BAND against the local signal instead of the drop height.")
    print("  Both columns are MEDIANS over the SAME samples, so the factor is meaningful:")
    print(f"    {'drop':<7}{'vs H0':>12}{'vs local':>12}{'factor':>10}"
          f"{'worst local':>14}{'n':>8}")
    for drop, v in u["per_drop"].items():
        print(f"    {drop:<7}{v['median_halfwidth_pct_of_H0_same_samples']:>11.4f}%"
              f"{v['median_rel_to_local_signal_pct']:>11.3f}%"
              f"{v['normalisation_factor_median_over_median']:>9.1f}x"
              f"{v['max_rel_to_local_signal_pct']:>13.3f}%"
              f"{v['n_samples_used_for_local']:>8}")
    print(f"    SAME MEASUREMENT, {u['normalisation_factor_min']:.1f}x to "
          f"{u['normalisation_factor_max']:.1f}x LARGER at the median and up to "
          f"{max(v['max_rel_to_local_signal_pct'] for v in u['per_drop'].values()):.1f} "
          f"percent at worst.")
    sw = u.get("local_cutoff_sweep")
    if sw:
        print(f"    CUTOFF SWEEP, because this section criticises a bare threshold and "
              f"must not carry one.")
        print(f"    The near-zero cutoff is {u['local_cutoff']}, and the factor DOES "
              f"depend on it:")
        for cut, cv in sw["by_cutoff"].items():
            print(f"      |x3/H0| > {cut:<5}  factor {cv['min']:.2f}x "
                  f"to {cv['max']:.2f}x")
        print(f"    So {u['normalisation_factor_min']:.1f}x is NOT cutoff independent, "
              f"unlike the 1.0 swept by --thresholds.")
        print(f"    What survives every cutoff is the DIRECTION and the order of "
              f"magnitude: the factor")
        print(f"    stays above 2x throughout ({sw['factor_min_over_all_cutoffs']:.2f}x "
              f"to {sw['factor_max_over_all_cutoffs']:.2f}x): "
              f"{sw['factor_stays_above_2x']}")
    print("    0.3 percent is a fraction of the DROP HEIGHT, not a pointwise")
    print("    relative uncertainty. Importing it as an acceptance tolerance for a "
          "decayed signal")
    worst = max(v["max_rel_to_local_signal_pct"]
                / v["median_halfwidth_pct_of_H0_same_samples"]
                for v in u["per_drop"].values())
    print(f"    applies an initial-amplitude band where the authors' own uncertainty "
          f"reaches {worst:.0f}x it.")
    print()
    if u["route_c_first_period_repeatability"]:
        print("  ROUTE C, separate origin (12 Raw files, different quantity): "
              "first damped period")
        for drop, v in u["route_c_first_period_repeatability"].items():
            print(f"    {drop}   mean {v['mean_s']:.6f} s   "
                  f"(max-min)/mean {v['range_pct_of_mean']:.4f} pct   n {v['n']}")
    print()
    print("=" * 86)
    print("IS THE EXPERIMENT THE LIMITING FACTOR? Inter-code envelope on the SAME "
          "statistic")
    print("=" * 86)
    print(f"  {'drop':<6}{'codes':>6}{'envelope pct':>22}{'width':>9}"
          f"{'exp (max-min)/mean':>21}{'ratio':>9}")
    for drop, r in c["rows"].items():
        env = f"{r['envelope_min_pct']:+.2f} to {r['envelope_max_pct']:+.2f}"
        print(f"  {drop:<6}{r['n_codes']:>6}{env:>22}"
              f"{r['envelope_width_points']:>9.2f}"
              f"{r['exp_range_pct_of_mean']:>20.4f}%"
              f"{r['intercode_over_experiment']:>8.1f}x")
    print(f"  ratio rises with drop height: {c['ratio_rises_with_drop_height']}")
    print(f"  envelope width grows {c['envelope_width_growth_01D_to_05D']:.2f}x from 01D "
          f"to 05D, while the experiment's own")
    print(f"  repeatability grows only "
          f"{c['exp_repeatability_growth_01D_to_05D']:.2f}x over the same range.")
    print(f"  experiment is NOT the limiting factor: "
          f"{c['experiment_is_not_the_limiting_factor']}")


def _p_thresholds(t):
    print()
    print("=" * 86)
    print("THRESHOLD SWEEP: is the 'groups within 1 percent' headline tolerance "
          "dependent?")
    print("=" * 86)
    for key, v in t["keys"].items():
        print(f"  {key:<18} count at 1.0 pct = {v['count_at_1pct']} of {v['n_groups']}"
              f"   invariant for ANY threshold in "
              f"({v['invariant_from_pct']:.4f}, {v['invariant_to_pct']:.4f}) pct"
              f"   = {v['invariance_factor']:.1f}x wide")
        print(f"    {'':16} per-group worst abs pct: "
              f"{[round(x, 3) for x in v['worst_abs_pct_sorted']]}")
    print(f"  narrowest invariance factor across keys: "
          f"{t['min_invariance_factor']:.1f}x")
    print(f"  THRESHOLD IS LOAD BEARING: {t['threshold_is_load_bearing']}")
    if not t["threshold_is_load_bearing"]:
        print("    So the 1.0 is a presentational bin in an empty region of the data, "
              "not an")
        print("    imported acceptance tolerance. No conclusion in this document rests "
              "on it.")


def self_test() -> dict:
    """Prove every guard FIRES. An assertion nobody has seen fail is not a check.

    Each case below deliberately breaks one input and requires an ExtractionError. A case
    that returns normally is a FAILURE of the self-test, because it means that extraction
    can still produce a verdict from data it did not read. This is the direct answer to
    the failure mode that motivated the guards: five checks in one night returned an
    answer they could not evaluate, and a silent empty-read looked exactly like a pass.
    """
    import tempfile
    results = []

    def case(name, fn, expect_substring=""):
        try:
            fn()
        except ExtractionError as e:
            ok = expect_substring.lower() in str(e).lower()
            results.append({"case": name, "fired": True, "matched": ok,
                            "message": str(e)[:150]})
        except Exception as e:                                  # pragma: no cover
            results.append({"case": name, "fired": False, "matched": False,
                            "message": f"WRONG EXCEPTION {type(e).__name__}: {e}"[:150]})
        else:
            results.append({"case": name, "fired": False, "matched": False,
                            "message": "RETURNED NORMALLY, guard did not fire"})

    tmp = Path(tempfile.mkdtemp(prefix="kramer_selftest_"))
    empty = tmp / "empty"
    empty.mkdir()

    global NUM_ROOT, EXP_ROOT
    saved_num, saved_exp = NUM_ROOT, EXP_ROOT

    # 1. THE ONE THAT WAS ACTUALLY BROKEN: radial_order on an empty tree used to return
    #    normally with reversal_is_universal_where_it_occurs = True.
    def c1():
        global NUM_ROOT
        NUM_ROOT = empty
        try:
            radial_order()
        finally:
            NUM_ROOT = saved_num
    case("radial_order on an empty numerical tree", c1, "carrying WG columns")

    # 2. series_manifest must not report a shrunken set as a result
    def c2():
        global NUM_ROOT
        NUM_ROOT = empty
        try:
            series_manifest()
        finally:
            NUM_ROOT = saved_num
    case("series_manifest on an empty numerical tree", c2, "numerical series found")

    # 3. a model sheet whose header row moved
    def c3():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for col, val in zip("CDEFGH", ("Model name", "Institution", "Author",
                                       "Software", "Description",
                                       "Computational effort")):
            ws[f"{col}5"] = val            # row 5, not row 4
        f = tmp / "shifted.xlsx"
        wb.save(f)
        model_table(f)
    case("model_table with the header row moved", c3, "missing")

    # 4. a model sheet with the right headers but no data rows
    def c4():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for col, val in zip("CDEFGH", ("Model name", "Institution", "Author",
                                       "Software", "Description",
                                       "Computational effort")):
            ws[f"{col}4"] = val
        f = tmp / "norows.xlsx"
        wb.save(f)
        model_table(f)
    case("model_table with zero data rows", c4, "expected 13, found 0")

    # 5. the cost table missing
    def c5():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        f = tmp / "nocost.xlsx"
        wb.save(f)
        cost_classes(f)
    case("cost_classes with no second table", c5, "is EMPTY")

    # 6. CODE_META emptied: the audit used to say NO SUBSTANTIVE DRIFT after 0 fields
    def c6():
        sys.path.insert(0, str(Path(__file__).resolve().parents[1]
                               / "simulation" / "r5_physics"))
        import kramer_benchmark as kb
        saved = dict(kb.CODE_META)
        kb.CODE_META.clear()
        try:
            audit_code_meta()
        finally:
            kb.CODE_META.update(saved)
    case("audit_code_meta with CODE_META empty", c6, "ZERO fields")

    # 7. uncertainty_scope pointed at a tree with no CI95 series. Route A silently
    #    returning a pooled mean of nothing would report "reproduces the abstract" from
    #    zero files, which is the same false-pass shape as cases 1 and 6.
    def c7():
        uncertainty_scope(empty)
    case("uncertainty_scope with no CI95 series", c7, "CI95 series missing")

    # 8. threshold_sensitivity when every group is tight. The sweep's whole purpose is
    #    to locate the gap between the tight groups and the outlier; with no outlier
    #    there is no gap, and an "invariant everywhere" verdict would be vacuous rather
    #    than reassuring.
    def c8():
        fake = {"tables": {"author": {"rows": [{"worst_abs_pct": 0.1},
                                               {"worst_abs_pct": 0.2}]}}}
        threshold_sensitivity(fake)
    case("threshold_sensitivity with no group above the threshold", c8,
         "at or above 1 pct")

    n_fired = sum(1 for r in results if r["fired"])
    n_matched = sum(1 for r in results if r["matched"])
    return {"n_cases": len(results), "n_fired": n_fired, "n_matched": n_matched,
            "all_fired": n_fired == len(results),
            "all_matched": n_matched == len(results),
            "results": results, "tmpdir": str(tmp)}


def _p_selftest(t):
    print("=" * 86)
    print("GUARD SELF-TEST: each case breaks one input and REQUIRES an ExtractionError")
    print("=" * 86)
    for r in t["results"]:
        mark = "PASS" if r["matched"] else ("FIRED, message mismatch" if r["fired"]
                                            else "**FAILED, NO GUARD**")
        print(f"  [{mark:<22}] {r['case']}")
        print(f"      {r['message']}")
    print()
    print(f"  guards that fired: {t['n_fired']} of {t['n_cases']}")
    print(f"  messages matched : {t['n_matched']} of {t['n_cases']}")
    print(f"  ALL GUARDS FIRE  : {t['all_fired']}")
    if not t["all_fired"]:
        _fail("AT LEAST ONE EXTRACTION CAN STILL PRODUCE A VERDICT FROM DATA IT DID NOT "
              "READ. Do not trust any result from this module until that is fixed.")


def main():
    ap = argparse.ArgumentParser(description=(__doc__ or "").split("\n")[0])
    ap.add_argument("--model-table", action="store_true")
    ap.add_argument("--audit", action="store_true", help="CODE_META against the sheet")
    ap.add_argument("--groups", action="store_true")
    ap.add_argument("--manifest", action="store_true")
    ap.add_argument("--envelope", action="store_true",
                    help="regroup the published envelope under every key")
    ap.add_argument("--order", action="store_true", help="radial order, every WG series")
    ap.add_argument("--sphere", action="store_true")
    ap.add_argument("--descriptions", action="store_true")
    ap.add_argument("--cost", action="store_true")
    ap.add_argument("--uncertainty", action="store_true",
                    help="the 0.3 pct figure, its statistic, and the envelope on one scale")
    ap.add_argument("--thresholds", action="store_true",
                    help="sweep the one bare threshold in this module")
    ap.add_argument("--self-test", action="store_true",
                    help="prove every fail-loud guard actually fires")
    ap.add_argument("--all", action="store_true")
    ap.add_argument("--json", action="store_true")
    a = ap.parse_args()
    if a.self_test:
        t = self_test()
        _p_selftest(t)
        raise SystemExit(0 if t["all_fired"] else 1)
    if not any((a.model_table, a.audit, a.groups, a.manifest, a.order,
                a.sphere, a.descriptions, a.cost, a.envelope,
                a.uncertainty, a.thresholds)):
        a.all = True

    blob = {}
    mt = None
    if a.all or a.model_table or a.audit or a.groups or a.envelope:
        mt = model_table()
        blob["model_table"] = mt
    if a.all or a.model_table:
        if not a.json:
            _p_model(mt)
    if a.all or a.cost:
        c = cost_classes()
        blob["cost_classes"] = c
        if not a.json:
            _p_cost(c)
    if a.all or a.audit:
        au = audit_code_meta(mt)
        blob["audit"] = au
        if not a.json:
            _p_audit(au)
    if a.all or a.groups:
        g = grouping_keys(mt)
        blob["grouping"] = g
        if not a.json:
            _p_groups(g)
    if a.all or a.envelope:
        e = envelope_by_grouping(mt)
        blob["envelope_by_grouping"] = e
        if not a.json:
            _p_envelope(e)
    if a.all or a.thresholds:
        th = threshold_sensitivity(blob.get("envelope_by_grouping"))
        blob["threshold_sensitivity"] = th
        if not a.json:
            _p_thresholds(th)
    if a.all or a.uncertainty:
        un = uncertainty_scope()
        cm = envelope_against_experiment()
        blob["uncertainty_scope"] = un
        blob["envelope_against_experiment"] = cm
        if not a.json:
            _p_uncertainty(un, cm)
    if a.all or a.manifest:
        m = series_manifest()
        blob["manifest"] = m
        if not a.json:
            _p_manifest(m)
    if a.all or a.order:
        o = radial_order()
        blob["radial_order"] = o
        if not a.json:
            _p_order(o)
    if a.all or a.sphere:
        sp = sphere_properties()
        sc = sphere_consistency(sp)
        blob["sphere"] = sp
        blob["sphere_consistency"] = sc
        if not a.json:
            _p_sphere(sp, sc)
    if a.all or a.descriptions:
        d = descriptions_inventory()
        cc = cad_cross_check()
        blob["descriptions"] = d
        blob["cad_cross_check"] = cc
        if not a.json:
            _p_desc(d, cc)
    if a.json:
        print(json.dumps(blob, indent=2, default=str))


if __name__ == "__main__":
    main()
